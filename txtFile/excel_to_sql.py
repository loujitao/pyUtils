#-*- coding: UTF-8 -*-
from openpyxl import load_workbook
'''
    脚本功能： 根据excel模板文档，自动生成hive建表语句sql文件
'''



'''
    data_path: excel文档的绝对路径
    index: excel文档的第几个sheet，从1开始计数 
'''
def read_sheet(data_path, index):
    wb = load_workbook(data_path, data_only=True)
    sheet_names = wb.sheetnames
    name = sheet_names[index-1]
    print("sheet: " + name)
    sheet = wb[name]
    #sheet = wb[sheet_name]
    return sheet

'''
    截取固定的表名
    sheet: excel的sheet表格对象
    col: 单元格坐标，格式如: "A1", "D19"
'''
def get_table_name(sheet, col):
    tables = sheet[col].value
    tables = tables.split("-")
    table_name = tables[0]
    table_name_zn = tables[1]
    print("table_name: " + table_name + ", table_name_cn: " + table_name_zn)
    return (table_name, table_name_zn)

'''
     start_row: 字段列起始行
     end_row：字段列结束行 
     col： 列英文名称
     col_cn：  列中文名称
     col_type： 字段列类型
'''
def get_table_col(sheet, start_row, end_row, col, col_cn, col_type):
    # 定义一个空列表
    datas = []
    j = start_row
    while j <= end_row:
        #print("j :" + str(j))
        column = sheet.cell(j, col).value
        column_cn = sheet.cell(j, col_cn).value
        column_type = sheet.cell(j, col_type).value
        j = j + 1
        if (column is None) or (column_cn is None):
            continue
        content = "{col_txt} {col_type_txt} COMMENT '{col_cn_txt}' ,".format(col_txt=column, col_type_txt=column_type,col_cn_txt=column_cn)
        # print(content)
        datas.append(content)
    return datas

'''
    生成hive建表的sql语句，并写入到文件中
'''
def write_sql(db_name,table_name, table_zn, cols, sql_path):
    sql = "create table if not exists " + db_name+"."+table_name + "(\n"
    for i in cols:
        sql += i + "\n"
    sql += ") COMMENT '" + table_zn + "' \n"
    sql += "PARTITIONED BY (dt STRING, ht STRING, mt STRING) \n"
    sql += "row format delimited fields terminated by ',' \n"
    sql += "STORED AS PARQUET \n"
    # 设置文件对象 a追加模式
    with open(sql_path, 'a') as f:
        f.write(sql + '\n\n')


if __name__ == '__main__':
    excel_path = "E:/11.xlsx"
    sheet = read_sheet(excel_path, 3)

    #生成SRC层的sql文件
    (src_table_name, src_table_name_zn) = get_table_name(sheet, "K1")
    sql_path = "E:/crt_{table_name}.sql".format(table_name=src_table_name)
    print("sql_path: " + sql_path)
    cols = get_table_col(sheet, 9, 84, 10, 11, 12)
    write_sql("src", src_table_name, src_table_name_zn, cols, sql_path)

    # 生成ODS层的sql文件
    (ods_table_name, ods_talbe_name_zn) = get_table_name(sheet, "S1")
    sql_path = "E:/crt_{table_name}.sql".format(table_name=ods_table_name)
    print("sql_path: " + sql_path)
    cols = get_table_col(sheet, 9, 84, 18, 19, 20)
    write_sql("ods", src_table_name, src_table_name_zn, cols, sql_path)
