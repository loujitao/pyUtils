from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from copy import copy

class Excel_data():
    def __init__(self, data_path, sheetname):
        #定义一个属性接收文件路径
        self.data_path = data_path
        # 定义一个属性接收工作表名称
        self.sheetname = sheetname
        # 打开指定工作簿
        self.wb = load_workbook(data_path)
        sheet_names = self.wb.sheetnames
        # 定位到指定的表单    workbook['表单名']
        #self.sheet = self.wb[sheetname]
        self.sheet = self.wb[sheet_names[1]]
        # 最大有效行
        self.max_row = self.sheet.max_row
        self.start_row = 9

    '''
        复制源头单元格样式，应用到目标单元格
    '''
    def copy_style(self, source_cell, target_cell):
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)
        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)
    '''
       补全表头
    '''
    def write_header(self):
        # SRC层表头信息
        self.sheet['J1'] = 'SRC表名（英文-中文）：'
        self.copy_style(self.sheet['E1'], self.sheet['J1'])
        self.sheet.merge_cells('K1:O1')
        self.sheet['J2'] = '设计人员'
        self.copy_style(self.sheet['E1'], self.sheet['J2'])
        self.sheet['K2'] = '娄继涛'
        self.sheet.merge_cells('K2:O2')
        self.sheet['J3'] = '版本-日期：'
        self.copy_style(self.sheet['E1'], self.sheet['J3'])
        self.sheet.merge_cells('K3:O3')
        self.sheet['J4'] = '分区方式:'
        self.copy_style(self.sheet['E1'], self.sheet['J4'])
        self.sheet['K4'] = ' dt=YYYY-MM-DD， ht=HH , mt=mm'
        self.sheet.merge_cells('K4:O4')
        self.sheet['J5'] = '存储组件：'
        self.copy_style(self.sheet['E1'], self.sheet['J5'])
        self.sheet['K5'] = 'HDFS  PARQUET'
        self.sheet.merge_cells('K5:O5')
        self.sheet['J6'] = 'SRC层字段明细'
        self.copy_style(self.sheet['A6'], self.sheet['J6'])
        self.sheet.merge_cells('J6:O6')
        # SRC层列名表头
        self.sheet['J7'] = '字段名'
        self.copy_style(self.sheet['A7'], self.sheet['J7'])
        self.sheet['K7'] = '字段中文名'
        self.copy_style(self.sheet['A7'], self.sheet['K7'])
        self.sheet['L7'] = '字段类型'
        self.copy_style(self.sheet['A7'], self.sheet['L7'])
        self.sheet['M7'] = '主键'
        self.copy_style(self.sheet['A7'], self.sheet['M7'])
        self.sheet['N7'] = '必填'
        self.copy_style(self.sheet['A7'], self.sheet['N7'])
        self.sheet['O7'] = '脱敏加密方式'
        self.copy_style(self.sheet['A7'], self.sheet['O7'])
        # SRC层说明行
        self.sheet['J8'] = '与源表尽量保持一致\n如源表缺失须补充'
        self.sheet['J8'].alignment = Alignment(wrapText=True)
        self.copy_style(self.sheet['F8'], self.sheet['J8'])
        self.sheet['K8'] = '与源表尽量保持一致\n如源表缺失须补充'
        self.sheet['K8'].alignment = Alignment(wrapText=True)
        self.copy_style(self.sheet['F8'], self.sheet['K8'])
        self.sheet['M8'] = '是/否'
        self.copy_style(self.sheet['F8'], self.sheet['M8'])
        self.sheet['N8'] = '是/否'
        self.copy_style(self.sheet['F8'], self.sheet['N8'])
        self.sheet['O8'] = '解密/加密算法\n（GMSM4|GMSM3）'
        self.sheet['O8'].alignment = Alignment(wrapText=True)
        self.copy_style(self.sheet['F8'], self.sheet['O8'])

        # ODS层表头信息
        self.sheet['P1'] = 'ODS表名（英文-中文）：'
        self.copy_style(self.sheet['A1'], self.sheet['P1'])
        self.sheet.merge_cells('P1:R1')
        self.sheet.merge_cells('S1:W1')
        self.sheet['P2'] = '设计人员'
        self.copy_style(self.sheet['A1'], self.sheet['P2'])
        self.sheet.merge_cells('P2:R2')
        self.sheet['S2'] = '娄继涛'
        self.sheet.merge_cells('S2:W2')
        self.sheet['P3'] = '版本-日期：'
        self.copy_style(self.sheet['A1'], self.sheet['P3'])
        self.sheet.merge_cells('P3:R3')
        self.sheet.merge_cells('S3:W3')
        self.sheet['P4'] = '分区方式:'
        self.copy_style(self.sheet['A1'], self.sheet['P4'])
        self.sheet.merge_cells('P4:R4')
        self.sheet['S4'] = ' dt=YYYY-MM-DD， ht=HH , mt=mm'
        self.sheet.merge_cells('S4:W4')
        self.sheet['P5'] = '存储组件：'
        self.copy_style(self.sheet['A1'], self.sheet['P5'])
        self.sheet.merge_cells('P5:R5')
        self.sheet['S5'] = 'HDFS  PARQUET'
        self.sheet.merge_cells('S5:W5')
        self.sheet['P6'] = 'ODS层字段明细'
        self.copy_style(self.sheet['A6'], self.sheet['P6'])
        self.sheet.merge_cells('P6:W6')
        # ODS层列名表头
        self.sheet['P7'] = '入仓'
        self.copy_style(self.sheet['A7'], self.sheet['P7'])
        self.sheet['Q7'] = '设计理由'
        self.copy_style(self.sheet['A7'], self.sheet['Q7'])
        self.sheet['R7'] = '字段名'
        self.copy_style(self.sheet['A7'], self.sheet['R7'])
        self.sheet['S7'] = '字段中文名'
        self.copy_style(self.sheet['A7'], self.sheet['S7'])
        self.sheet['T7'] = '字段类型'
        self.copy_style(self.sheet['A7'], self.sheet['T7'])
        self.sheet['U7'] = '主键'
        self.copy_style(self.sheet['A7'], self.sheet['U7'])
        self.sheet['V7'] = '必填'
        self.copy_style(self.sheet['A7'], self.sheet['V7'])
        self.sheet['W7'] = '数据清洗规则'
        self.copy_style(self.sheet['A7'], self.sheet['W7'])
        # ODS层说明行
        self.sheet['P8'] = '是/否'
        self.copy_style(self.sheet['F8'], self.sheet['P8'])
        self.sheet['Q8'] = '描述是否入仓的理由'
        self.copy_style(self.sheet['F8'], self.sheet['Q8'])
        self.sheet['R8'] = '建议优化字段名称'
        self.copy_style(self.sheet['F8'], self.sheet['R8'])
        self.sheet['S8'] = '建议优化字段名称'
        self.copy_style(self.sheet['F8'], self.sheet['S8'])
        self.sheet['U8'] = '是/否'
        self.copy_style(self.sheet['F8'], self.sheet['U8'])
        self.sheet['V8'] = '是/否'
        self.copy_style(self.sheet['F8'], self.sheet['V8'])
        self.sheet['W8'] = '一般包含转换数据格式和统一维度'
        self.copy_style(self.sheet['F8'], self.sheet['W8'])

    '''
      补全SRC层的内容  
    '''
    def write_SRC(self):
        i = self.start_row
        while i <= self.max_row:
            #print('SRC: excel第{no}行数据！'.format(no=i))
            #  B列 英文字段  -》 J列 英文字段
            self.sheet.cell(i, 10).value = self.sheet.cell(i, 2).value
            self.copy_style(self.sheet.cell(i, 2), self.sheet.cell(i, 10))
            # 中文字段  -》  SRC中文字段
            self.sheet.cell(i, 11).value = self.sheet.cell(i, 3).value
            self.copy_style(self.sheet.cell(i, 3), self.sheet.cell(i, 11))
            self.sheet.cell(i, 12).value = 'string'
            self.copy_style(self.sheet.cell(i, 4), self.sheet.cell(i, 12))
            self.sheet.cell(i, 13).value = self.sheet.cell(i, 6).value
            self.copy_style(self.sheet.cell(i, 6), self.sheet.cell(i, 13))
            self.sheet.cell(i, 14).value = self.sheet.cell(i, 7).value
            self.copy_style(self.sheet.cell(i, 7), self.sheet.cell(i, 14))
            self.sheet.cell(i, 15).value = self.sheet.cell(i, 9).value
            self.copy_style(self.sheet.cell(i, 9), self.sheet.cell(i, 15))
            i = i + 1

    '''
      补全ODS层的内容  
    '''
    def write_ODS(self):
        i = self.start_row
        while i <= self.max_row:
            #print('ODS: excel第{no}行数据！'.format(no=i))
            #  B列 英文字段  -》 J列 英文字段
            self.sheet.cell(i, 16).value = '是'
            self.copy_style(self.sheet.cell(i, 2), self.sheet.cell(i, 16))
            self.sheet.cell(i, 17).value = self.sheet.cell(i, 3).value
            self.copy_style(self.sheet.cell(i, 2), self.sheet.cell(i, 17))
            lower = '=LOWER(B{number})'.format(number=i)
            self.sheet.cell(i, 18).value = lower
            self.copy_style(self.sheet.cell(i, 2), self.sheet.cell(i, 18))
            self.sheet.cell(i, 19).value = self.sheet.cell(i, 3).value
            self.copy_style(self.sheet.cell(i, 2), self.sheet.cell(i, 19))
            self.sheet.cell(i, 20).value = 'string'
            self.copy_style(self.sheet.cell(i, 2), self.sheet.cell(i, 20))
            self.sheet.cell(i, 21).value = self.sheet.cell(i, 6).value
            self.copy_style(self.sheet.cell(i, 2), self.sheet.cell(i, 21))
            self.sheet.cell(i, 22).value = self.sheet.cell(i, 7).value
            self.copy_style(self.sheet.cell(i, 2), self.sheet.cell(i, 22))
            self.copy_style(self.sheet.cell(i, 2), self.sheet.cell(i, 23))
            i = i + 1

    '''
       保存excel文档，最后执行一次
    '''
    def save_excel(self):
        self.wb.save(self.data_path)



path = "E:/xuqiu.xlsx"
sheet_name = "THIRD_ORDER_INFO"
sheet = Excel_data(path, sheet_name)
sheet.write_header()
sheet.write_SRC()
sheet.write_ODS()
sheet.save_excel()