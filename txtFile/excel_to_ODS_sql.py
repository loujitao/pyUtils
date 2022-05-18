#-*- coding: UTF-8 -*-
from openpyxl import load_workbook
'''
    脚本功能： 根据excel模板文档，自动生成hive建表语句sql文件
'''



'''
    data_path: excel文档的绝对路径
    index: excel文档的第几个sheet，从1开始计数 
'''
def read_sheet(data_path, index):
    wb = load_workbook(data_path, data_only=True)
    sheet_names = wb.sheetnames
    name = sheet_names[index-1]
    print("sheet: " + name)
    sheet = wb[name]
    #sheet = wb[sheet_name]
    return sheet

'''
    截取固定的表名
    sheet: excel的sheet表格对象
    col: 单元格坐标，格式如: "A1", "D19"
'''
def get_table_name(sheet, col):
    tables = sheet[col].value
    tables = tables.split("-")
    table_name = tables[0]
    table_name_zn = tables[1]
    print("table_name: " + table_name + ", table_name_cn: " + table_name_zn)
    return (table_name, table_name_zn)

'''
     start_row: 字段列起始行
     end_row：字段列结束行 
     col： 列英文名称
     col_cn：  列中文名称
     col_type： 字段列类型
'''
def get_table_col(sheet, start_row, end_row, col, col_cn, col_type):
    # 定义一个空列表
    datas = []
    j = start_row
    while j <= end_row:
        #print("j :" + str(j))
        column = sheet.cell(j, col).value
        column_cn = sheet.cell(j, col_cn).value
        column_type = sheet.cell(j, col_type).value
        j = j + 1
        if (column is None) or (column_cn is None):
            continue
        content = "{col_txt}  {col_type_txt}  COMMENT  '{col_cn_txt}' ,".format(col_txt=column, col_type_txt=column_type,col_cn_txt=column_cn)
        # print(content)
        datas.append(content)
    return datas

'''
    生成hive建表的sql语句，并写入到文件中
'''
def write_sql(db_name,table_name, table_zn, cols, sql_path):
    sql = "create table if not exists " + db_name+"."+table_name + "(\n"
    for i in cols:
        sql += "  " + i + "\n"
    sql += ") COMMENT '" + table_zn + "' \n"
    sql += "PARTITIONED BY (dt STRING, ht STRING, mt STRING) \n"
    sql += "STORED AS PARQUET \n"
    sql += "TBLPROPERTIES ( \n"
    sql += "    'partition.time-extractor.kind' = 'custom',\n"
    sql += "    'partition.time-extractor.class' = 'com.epay.ts.fdw.flink.func.util.CustomWaterMarkTimeExtractor',\n"
    sql += "    'partition.time-extractor.timestamp-pattern' = '$dt $ht:$mt:00',\n"
    sql += "    'sink.partition-commit.trigger' = 'partition-time',\n"
    sql += "    'sink.partition-commit.delay' = '0 s',\n"
    sql += "    'sink.parallelism' = '1',\n"
    sql += "    'sink.partition-commit.policy.kind' = 'metastore,success-file,custom',\n"
    sql += "    'sink.partition-commit.policy.class' = 'com.epay.ts.fdw.flink.func.util.ParquetFileMergingCommitPolicy'\n"
    sql += " ); \n"
    # 设置文件对象 a追加模式
    with open(sql_path, 'a') as f:
        f.write(sql + '\n\n')


def get_insert_col(sheet, start_row, end_row, src_col, ods_col ):
    # 定义一个空列表
    datas = []
    j = start_row
    while j <= end_row:
        #print("j :" + str(j))
        src_column = sheet.cell(j, src_col).value
        ods_column = sheet.cell(j, ods_col).value
        j = j + 1
        if (ods_column is None):
            continue
        content = "{src_column_txt}  as  {ods_column_txt} ,".format(src_column_txt=src_column, ods_column_txt=ods_column)
        datas.append(content)
    return datas

def write_insert(db_name, ods_table_name, cols, src_table_name,  sql_path):
    sql = "INSERT OVERWRITE TABLE "+ db_name+"."+ ods_table_name + " PARTITION (dt, ht, mt) \n"
    sql += "select \n"
    for i in cols:
        sql += "  " + i + "\n"
    sql += " '$v_date'  AS dt, \n"
    sql += " '23'  AS ht, \n"
    sql += " '60'  AS mt \n"
    sql += "FROM src." + src_table_name + " \n"
    sql += "WHERE dt = FROM_UNIXTIME(UNIX_TIMESTAMP('$v_date', 'yyyy-MM-dd'), 'yyyy-MM-dd'); \n"
    sql += "\n"
    sql += "ALTER TABLE "+ db_name+"."+ ods_table_name +" DROP IF EXISTS PARTITION (dt = '$v_date',ht<='23',mt<>'60');"
    # 设置文件对象 a追加模式
    with open(sql_path, 'a') as f:
        f.write(sql + '\n')


if __name__ == '__main__':
    dir_path = r"D:\gitDoc\ts-fdw-doc\数仓设计和开发\1.SRC&ODS层\沃分期系统-ISIM/"
    file_name = "THIRD_ORDER_INFO-沃分期3.0订单表.xlsx"

    excel_path = dir_path + file_name
    print("excel_path" + excel_path)
    sheet = read_sheet(excel_path, 3)

    # 生成ODS层的sql文件
    (ods_table_name, ods_talbe_name_zn) = get_table_name(sheet, "S1")
    sql_path = "E:/aaaa/{table_name}.sql".format(table_name= ods_table_name.lower())
    print("sql_path: " + sql_path)
    ##### get_table_col(sheet, start_row, end_row, col, col_cn, col_type) 后三个参数基本固定
    cols = get_table_col(sheet, 9, 200, 18, 19, 20)
    write_sql("ods", ods_table_name, ods_talbe_name_zn, cols, sql_path)

    ### get_insert_col(sheet, start_row, end_row, src_col, ods_col)  后两个参数基本固定
    insert_cols = get_insert_col(sheet, 9, 200, 10, 18)
    (src_table_name, src_table_name_zn) = get_table_name(sheet, "K1")
    write_insert("ods", ods_table_name, insert_cols, src_table_name, sql_path)


